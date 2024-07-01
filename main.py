from datetime import datetime
import reportlab
from reportlab.pdfgen import canvas
from PIL import Image, ImageTk
from reportlab.platypus import Table, TableStyle, Image
from reportlab.lib.pagesizes import A5, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
import cv2
import tkinter
from tkinter import *
from tkinter import simpledialog
import time
import pymongo
from pdf2image import convert_from_path
from reportlab.pdfbase import pdfmetrics #for importing Hindi Font
from reportlab.pdfbase.ttfonts import TTFont #for importing Hindi Font
from tkinter import ttk # for importing treeview
from datetime import datetime, timedelta
from tkcalendar import DateEntry  # Tkinter Calendar
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import tkcalendar
from tkcalendar import Calendar, DateEntry
from openpyxl import Workbook

##########################################
# NOTES: THIS IS Atithi2.1 VER 1.5 under dev
# isInside and pass issue and return pass function working with updating DB.
# Sr no on passes and sessions
# reports and settings page are now working, they can have better UI but are functional
# addition of isoDate to the database to make the search effective with dates.

# PENDING
# digitalClock error happened after I introduced the reports function
# Create DB with script during installation
# Setting Location on footer and pdfs using information from setting page
# licensing
# distribution

#### Things for Windows
import os

############## Under progress tasks and objectives

global path
# print(path)
path=str(os.getcwd()).replace("\\","/")
poppler_path=r'C:\Program Files\poppler-0.68.0\bin'

#path to Hindi Font
font_name="gargi"

pdfmetrics.registerFont(TTFont('Hindi1', path+'/fonts/'+font_name+'.ttf'))


window=Tk()
window.title("IOCL Atithi2.1.1 by Egis Engineering")
window.geometry('990x600+240+70')
icon=PhotoImage(path+"/arriver.ico")
window.iconbitmap(icon)
window.configure(bg="white")


#logos
atithi_img=PhotoImage(file=path+'/images/arriver.png')
iocl_img=PhotoImage(file=path+'/images/2_iocl.png')

#Connection to MongoDB
uri = 'mongodb://127.0.0.1:27017/'
client = pymongo.MongoClient(uri)
database = client['atithi2']
visitors = database['visitors']
sessions=database['sessions']
officers_db=database["officers"]
location_db=database['location']
users_db = database['users']  # Users collection

#### Pass Count
availableCards=int((location_db.find({}, {"allowedVisitors": 1,"_id":0}))[0]["allowedVisitors"])
slots=[]
for i in range(availableCards):
    slots.append("V"+str(i+1))
# check what cards are assigned right now
def reviewCardAssignment():
    cardsAssigned=[]
    cardsAssignedfromDB = sessions.find({'isInside': 'yes'}, {'_id': 0, 'passNo': 1})
    # print(sessions.count_documents({'isInside':'yes'}))
    for i in cardsAssignedfromDB:
        # print(i)
        cardsAssigned.append(i['passNo'])

    global availableSlots
    availableSlots = [x for x in slots if x not in cardsAssigned]


header = Frame(window, bg='white')
home = Frame(window, bg='white')
newVisitor = Frame(window, bg='white')
findVisitor = Frame(window, bg='white')
frame_returnPass = Frame(window, bg='white')
footer = Frame(window, bg='black')


def clickPicture():
    global img_name
    global img_name_thumbnail
    now = time.strftime("%d%m%Y_%H%M%S")
    cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)  # captureDevice = camera for windows
    # cam = cv2.VideoCapture(1)
    cv2.namedWindow("Press Space to Click")
    # img_counter = 0
    while True:
        ret, frame = cam.read()
        if not ret:
            messagebox.showerror("Error", "failed to grab frame")
            break
        cv2.imshow("Press Space to Ckick", frame)

        k = cv2.waitKey(1)
        if k % 256 == 27:
            # ESC pressed
            # messagebox.showinfo("Closing", "Escape hit, closing...")
            break
        elif k % 256 == 32:
            # SPACE pressed
            img_name = path+"/visitorImages/"+ now +".png"
            # print(img_name)
            cv2.imwrite(img_name, frame)
            # messagebox.showinfo("success", "Image Clicked")
            #Size the image
            src = cv2.imread(img_name, cv2.IMREAD_UNCHANGED)
            # percent by which the image is resized
            scale_percent = 15
            # calculate the 50 percent of original dimensions
            width = int(src.shape[1] * scale_percent / 100)
            height = int(src.shape[0] * scale_percent / 100)
            # dsize
            dsize = (width, height)
            # resize image
            output = cv2.resize(src, dsize)
            img_name_thumbnail=path+"/visitorImages/visitorThumbnails/th_"+now + ".png"
            cv2.imwrite(img_name_thumbnail, output)
            break
            # print("{} written!".format(img_name))
            # img_counter += 1
    cam.release()
    cv2.destroyAllWindows()


def hide_all_frames():
    home.grid_forget()
    newVisitor.grid_forget()
    findVisitor.grid_forget()
    frame_returnPass.grid_forget()

def click_home():
    hide_all_frames()
    home.grid(row=1, sticky='news')
    generateHomeContent()


def click_reset():
    requirement =['firstName', 'lastName', 'mobileNumber','photoId', 'address','areaAllowed', 'officerVisited','purpose']
    inputs = [firstName_in, lastName_in, mobileNumber_in, photoId_in, address_in]
    firstName_in.delete(0,END)
    lastName_in.delete(0, END)
    mobileNumber_in.delete(0,END)
    photoId_in.delete(0,END)
    clickedId.set("Govt Photo ID")
    clickedOfficers.set("Officer Name")
    areaAllowed.set("General Area")
    address_in.delete(0,END)
    purpose_in.delete(0,END)

def submit():
    reviewCardAssignment()
    requirement = ['firstName', 'lastName', 'mobileNumber', 'photoId', 'address', 'areaAllowed','officerVisited', 'purpose']
    inputs = [firstName_in, lastName_in, mobileNumber_in, photoId_in, address_in, purpose_in]
    if len(firstName_in.get()) < 1 or len(lastName_in.get()) < 1 or len(mobileNumber_in.get()) < 1 or len(
            address_in.get()) < 1 or len(photoId_in.get()) < 1 or len(purpose_in.get()) < 1:
        messagebox.showerror("Error", "Fields can not be left blank")
        # click_reset()
    else:
        # print(img_name)
        firstName = str(firstName_in.get()).capitalize().strip(" ")
        lastName = str(lastName_in.get()).capitalize().strip(" ")
        mobileNumber = str(mobileNumber_in.get()).strip(" ")
        photoId = str(photoId_in.get()).upper().strip(" ")
        address = str(address_in.get()).capitalize()
        areaSelected=str(areaAllowed.get())
        officerName = str(clickedOfficers.get())
        purpose = str(purpose_in.get()).capitalize()
        docType = clickedId.get()
        verify = visitors.count_documents({'mobileNumber': mobileNumber})
        if verify > 0:
            messagebox.showerror("Error", "Number already exists")
        else:
            clickPicture()
            now = time.strftime(("%d%m%Y.%H%M%S"))

            sessionIdNow = (sessions.find({}, {"sessionId": 1, "_id": 0}).sort("_id", -1)[0]["sessionId"])
            if int(sessionIdNow) < 1:
                srNo = 1
            else:
                srNo = str(int(sessionIdNow) + 1)

            ########################################################

            # Function to convert string to ISODate
            def convert_to_isodate(date_str, format):
                try:
                    return datetime.strptime(date_str, format)
                except ValueError:
                    return None

            # Convert date fields to ISODate
            today = time.strftime("%d/%m/%Y")
            isoDate = convert_to_isodate(today, "%d/%m/%Y")




            ##########################################################

            visitors.insert_one({"timestamp": now,
                                 "firstName": firstName,
                                 "lastName": lastName,
                                 "mobileNumber": mobileNumber,
                                 "photoId": docType + "/" + photoId,
                                 "fullAddress": address,
                                 "visitorImage": img_name,
                                 "visitorThumbnail":img_name_thumbnail,
                                 # "areaAllowed":areaSelected,
                                 })
            sessions.insert_one({"timestamp": now,
                                 "firstName": firstName,
                                 "lastName": lastName,
                                 "mobileNumber": mobileNumber,
                                 "photoId": docType + "/" + photoId,
                                 "areaAllowed": areaSelected,
                                 "officerVisited": officerName,
                                 "Purpose": purpose,
                                 "entryDateTime": time.strftime("%d/%m/%Y-%H:%M:%S"),
                                 "entryDate":time.strftime("%d/%m/%Y"),
                                 "isoDate":isoDate,
                                 # "passNumber":entryPass,
                                 "isInside":"yes",
                                 "passNo":str(availableSlots[0]),
                                 "sessionId":srNo
                                 })
            messagebox.showinfo("Success", "Visitor enrolled successfully")
            # Begining of the PDF generation part
            now = datetime.now()
            datedSmall = now.strftime("%d,%h,%Y")
            dated = now.strftime("%A %d/%m/%Y")
            timed = now.strftime("%H:%M")

            # PDF Part
            def createPass():

                # Declare the fileName
                width, height = A5
                # Declare the fileName
                global entryPass
                entryPass = str(path+"/visitorPasses/" + firstName + lastName +datedSmall+ ".pdf")
                # print(entryPass)
                file = canvas.Canvas(entryPass, pagesize=landscape(A5))
                # file=SimpleDocTemplate("accessImages/PDFs/able.pdf")
                file.setTitle("Visitor Pass")
                logoPath = path+"/images/5_iocl.png"
                logo = reportlab.platypus.Image(logoPath)
                # logo = PhotoImage(file=logoPath)
                visitorImagePath = img_name_thumbnail
                visitorImage = reportlab.platypus.Image(visitorImagePath)
                # visitorImage = PhotoImage(file=visitorImagePath)


                #### adding data to the pass specific to the location using the DB

                data = [[logo, "Visitor Pass #"+srNo+"\n \n आगंतुक पहचान पास क #"+srNo, visitorImage],
                        ["", locationType_query+"\n \n इंडेन बॉटलिंग प्लांट", ""],
                        [ "Date: "+datedSmall+"\n" + "Time: "+timed, "Baddi, Himachal Pradesh\nबद्दी,हिमाचल प्रदेश", areaSelected],
                        ["Visitor Name \nआगंतुक का नाम", firstName + " " + lastName, ""],
                        ["Visitor Address\nआगंतुक का पता", address, ""],
                        ["Mobile Number\nमोबाइल नंबर", mobileNumber, ""],
                        ["Officer Name\nअधिकारी का नाम", officerName, ""],
                        ["Officer Signature\nअधिकारी हस्ताक्षर", "Gaurd Signature\nसुरक्षा अधिकारी हस्ताक्षर", "Visitor Signature\nआगंतुक हस्ताक्षर"],
                        ["", "", ""],
                        ["", "", ""],
                        ["", "", ""],
                        ]
                t = Table(data, colWidths=[2.5 * inch])
                # t=Table(data)
                ts = TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black,),
                    ('SPAN', (0, 0), (0, 1)),  # for the logo
                    ('SPAN', (2, 0), (2, 1)),  # for the Visitor Image
                    ('SPAN', (1, 3), (2, 3)),  # for Visitor Name
                    # ('ALIGN',(0, 2), (2,2)), #for QR code Alignment
                    ('ALIGN', (0, 0), (2, 1), 'CENTER'), #center align for top 3 rows
                    ('ALIGN', (2, 2), (2, 2), 'CENTER'),  # center align for licenced area cell
                    ('FONTSIZE', (2, 2), (2, 2), 15),  # Size of licensed area cell
                    ('VALIGN', (2, 2), (2, 2), 'MIDDLE'),  # V Alignment of licensed area cell
                    ('SPAN', (1, 3), (2, 3)),  # visitor infomration
                    ('SPAN', (1, 4), (2, 4)),  # visitor infomration
                    ('SPAN', (1, 5), (2, 5)),  # visitor infomration
                    ('SPAN', (1, 6), (2, 6)),  # visitor infomration
                    ('FONTSIZE',(0,0),(2,1),15), #Size of headers
                    ('FONT', (0, 0), (-1, -1), "Hindi1"),
                    ('VALIGN', (0, 0), (1, 1), 'TOP'),  # V Alignment of header
                    # ('ALIGN', (0, 1), (1, 1), 'CENTER'),  # V Alignment of header
                    ('SPAN', (0, 8), (0, 10)),  # for signature space
                    ('SPAN', (1, 8), (1, 10)),  # for signature space
                    ('SPAN', (2, 8), (2, 9)),  # for signature space

                    # ('BOTTOMPADDING',(-1,-1),(-1,-1),12)
                ])
                t.setStyle(ts)
                t.wrapOn(file, width, height)
                t.drawOn(file, 15 * mm, 20 * mm)

                elements = []
                elements.append(t)

                # file.build(elements)
                file.save()
            createPass()
############Create Document viewer feature
            def viewPass():
                def quit():
                    viewer.destroy()
                    click_reset()
                    click_home()
                def printIt():
                    click_reset()
                    viewer.destroy()
                    os.startfile(entryPass, "print") #### Only for Windows

                    click_home()
                # Create a root window
                viewer = tkinter.Toplevel()
                viewer.title("Visitor Pass Preview")
                # Set the width and height of our root window.
                viewer.geometry("500x450+400+200")
                viewer.config(bg="white")

                pdf_frame = Frame(master=viewer, bg='white').place(x=0,y=10)
                b_print = Button(viewer, text="Print", bg='orange', width=10, command=printIt)
                b_exit = Button(viewer, text="Exit", bg='orange', width=10, command=quit)
                b_print.place(x=120, y=5)
                b_exit.place(x=300, y=5)
                pdf = Text(viewer, bg="white")
                pdf.place(x=0, y=30)
                # Here the PDF is converted to list of images
                pages = convert_from_path(entryPass, size=(500, 440),poppler_path=poppler_path)
                # Empty list for storing images
                photos = []
                # Storing the converted images into list
                for i in range(len(pages)):
                    photos.append(ImageTk.PhotoImage(pages[i]))
                # Adding all the images to the text widget
                for photo in photos:
                    pdf.image_create(END, image=photo)

                    # For Seperating the pages
                    pdf.insert(END, '\n\n')
                # Ending of mainloop
                mainloop()
            viewPass()

def click_newVisitor():
    newVisitor.grid(row=1, sticky='news')
    # b_clickPicture.place(x=700, y=200)
    firstName.place(x=70, y=0)
    lastName.place(x=70, y=35)
    mobileNumber.place(x=70, y=70)
    photoId.place(x=70, y=105)
    address.place(x=70, y=145)
    area.place(x=70,y=175)
    officerVisited.place(x=70, y=210)
    purpose.place(x=70, y=245)

    firstName_in.place(x=370, y=0)
    lastName_in.place(x=370, y=35)
    mobileNumber_in.place(x=370, y=70)
    photoId_in.place(x=370, y=105)
    address_in.place(x=370, y=145)
    areaAllowed_in.place(x=370,y=175)
    officerVisited_in.place(x=370,y=210)
    purpose_in.place(x=370, y=245)

    reset.place(x=200, y=280)
    submit.place(x=400, y=280)
    b_home.place(x=600, y=280)


######## FIND VISITOR CODE #############
def openFindPage():
    global keysInDb, labelsForVisitorInfo
    labelsForVisitorInfo=['First Name','Last Name','Mobile Number', 'Photo Id','Address','visitorImage']
    keysInDb=["firstName",'lastName','mobileNumber','photoId','fullAddress','visitorThumbnail']
    def searchVisitorButton():
        numberToSearch=entry_find.get()
        #validate entry in search box
        if len(numberToSearch)<1:
            messagebox.showerror("Error","Value Can't Be Empty")
        else:
            if visitors.count_documents({"mobileNumber": numberToSearch})==0:
                messagebox.showerror("Error","No Reults Found")
            else:
                query=visitors.find({'mobileNumber':numberToSearch},{'_id':0,'timestamp':0})
                reply=query[0]
                #All the values returned from the DB about the visitor
                returnedVisitorInfo=[]
                # get values from DB
                for i in range(len(keysInDb)):
                    returnedVisitorInfo.append(reply[keysInDb[i]])
                dataOfVisitor=[keysInDb,returnedVisitorInfo]


                # CREATE TABLE
                def createTable():
                    table=Frame(findVisitor)
                    table.place(x=40,y=90)
                    for i in range(len(labelsForVisitorInfo)-1):
                        e=Entry(master=table,relief=GROOVE,bg='white')
                        e.grid(row=0,column=i)
                        e.insert(END,labelsForVisitorInfo[i])
                    for j in range(len(returnedVisitorInfo)-1):
                        e = Entry(master=table, relief=GROOVE,bg='white')
                        e.grid(row=1, column=j)
                        e.insert(END, returnedVisitorInfo[j])

                    def button_selectVisitor():
                        def createPass():

                            sessionIdNow = (
                            sessions.find({}, {"sessionId": 1, "_id": 0}).sort("_id", -1)[0]["sessionId"])
                            if int(sessionIdNow) < 1:
                                srNo = 1
                            else:
                                srNo = str(int(sessionIdNow) + 1)

                            def Ret_quit():
                                viewer.destroy()
                                visitorInfo.destroy()
                                click_home()

                            def Ret_printIt():
                                os.startfile(ret_entryPass, "print") #### Only for Windows
                                viewer.destroy()
                                visitorInfo.destroy()
                                click_home()
                                ########Submit Details to Session DB
                                push={}
                                ts = {"timestamp": now_time}
                                push.update(ts)

                                for i in range(len(keysInDb)):
                                    data = {keysInDb[i]:returnedVisitorInfo[i]}
                                    push.update(data)
                                reviewCardAssignment()

                                # Function to convert string to ISODate
                                def convert_to_isodate(date_str, format):
                                    try:
                                        return datetime.strptime(date_str, format)
                                    except ValueError:
                                        return None

                                # Convert date fields to ISODate
                                today = time.strftime("%d/%m/%Y")
                                isoDate = convert_to_isodate(today, "%d/%m/%Y")

                                otherDetails={"officerVisited":officerName,"areaAllowed":ret_areaSelected,"Purpose":purpose,"entryDateTime":time.strftime("%d/%m/%Y-%H:%M:%S"),"entryDate":time.strftime("%d/%m/%Y"),"isoDate":isoDate,
                                              "passNo":str(availableSlots[0]),"isInside":"yes","sessionId":srNo}
                                push.update(otherDetails)
                                sessions.insert_one(push)
                                messagebox.showinfo("Pass successful", "Visitor Pass Generated successfully")


                            # Data formats for timestamps
                            now = datetime.now()
                            datedSmall = now.strftime("%d,%h,%Y")
                            dated = now.strftime("%A %d/%m/%Y")
                            timed = now.strftime("%H:%M")
                            now_time = time.strftime("%d%m%Y.%H%M%S")
                            # now = datetime.now()
                            # datedSmall = now.strftime("%d,%h,%Y")
                            # dated = now.strftime("%A %d/%m/%Y")
                            # timed = now.strftime("%H:%M")


                            # Declare the fileName
                            width, height = A5
                            # Declare the fileName
                            global ret_entryPass
                            ret_entryPass = str(path+"/visitorPasses/" + returnedVisitorInfo[0] + returnedVisitorInfo[
                                1] + datedSmall + ".pdf")
                            file = canvas.Canvas(ret_entryPass, pagesize=landscape(A5))
                            file.setTitle("Visitor Pass")
                            logoPath = path+"/images/5_iocl.png"
                            logo = reportlab.platypus.Image(logoPath)
                            visitorImagePath = returnedVisitorInfo[5]
                            visitorImage = reportlab.platypus.Image(visitorImagePath)
                            officerName = ret_clickedOfficer.get()
                            ret_areaSelected=str(ret_areaAllowed.get())
                            purpose = str(ret_purpose_in.get()).capitalize()

                            data = [[logo, "Visitor Pass #"+srNo+"\n \n आगंतुक पहचान पास क #"+srNo, visitorImage],
                                    ["", "Indane Bottling Plant\n \n इंडेन बॉटलिंग प्लांट", ""],
                                    ["Date: " + datedSmall + "\n" + "Time: " + timed,"Baddi, Himachal Pradesh\nबद्दी,हिमाचल प्रदेश", ret_areaSelected],
                                    ["Visitor Name \nआगंतुक का नाम", returnedVisitorInfo[0] + " " + returnedVisitorInfo[1],""],
                                    ["Visitor Address\nआगंतुक का पता", returnedVisitorInfo[4], ""],
                                    ["Mobile Number\nमोबाइल नंबर",  returnedVisitorInfo[2], ""],
                                    ["Officer Name\nअधिकारी का नाम", officerName, ""],
                                    ["Officer Signature\nअधिकारी हस्ताक्षर","Gaurd Signature\nसुरक्षा अधिकारी हस्ताक्षर","Visitor Signature\nआगंतुक हस्ताक्षर"],
                                    ["", "", ""],
                                    ["", "", ""],
                                    ["", "", ""],
                                    ]
                            t = Table(data, colWidths=[2.5 * inch])
                            # t=Table(data)
                            ts = TableStyle([
                                ('GRID', (0, 0), (-1, -1), 1, colors.black,),
                                ('SPAN', (0, 0), (0, 1)),  # for the logo
                                ('SPAN', (2, 0), (2, 1)),  # for the Visitor Image
                                ('SPAN', (1, 3), (2, 3)),  # for Visitor Name
                                # ('ALIGN',(0, 2), (2,2)), #for QR code Alignment
                                ('ALIGN', (0, 0), (2, 1), 'CENTER'),  # center align for top 3 rows
                                ('ALIGN', (2, 2), (2, 2), 'CENTER'),  # center align for licenced area cell
                                ('FONTSIZE', (2, 2), (2, 2), 15),  # Size of licensed area cell
                                ('VALIGN', (2, 2), (2, 2), 'MIDDLE'),  # V Alignment of licensed area cell
                                ('SPAN', (1, 3), (2, 3)),  # visitor infomration
                                ('SPAN', (1, 4), (2, 4)),  # visitor infomration
                                ('SPAN', (1, 5), (2, 5)),  # visitor infomration
                                ('SPAN', (1, 6), (2, 6)),  # visitor infomration
                                ('FONTSIZE', (0, 0), (2, 1), 15),  # Size of headers
                                ('FONT', (0, 0), (-1, -1), "Hindi1"),
                                ('VALIGN', (0, 0), (1, 1), 'TOP'),  # V Alignment of header
                                # ('ALIGN', (0, 1), (1, 1), 'CENTER'),  # V Alignment of header
                                ('SPAN', (0, 8), (0, 10)),  # for signature space
                                ('SPAN', (1, 8), (1, 10)),  # for signature space
                                ('SPAN', (2, 8), (2, 9)),  # for signature space

                                # ('BOTTOMPADDING',(-1,-1),(-1,-1),12)
                            ])
                            t.setStyle(ts)
                            t.wrapOn(file, width, height)
                            t.drawOn(file, 15 * mm, 20 * mm)

                            elements = []
                            elements.append(t)
                            # file.build(elements)
                            file.save()
                            #preview the pass before printing
                            # Create a root window
                            viewer = tkinter.Toplevel()
                            viewer.title("Visitor Pass Preview")
                            # Set the width and height of our root window.
                            viewer.geometry("500x450+400+200")
                            viewer.config(bg="white")
                            pdf_frame = Frame(master=viewer, bg='white').place(x=0, y=10)
                            b_print = Button(viewer, text="Print", bg='orange', width=10, command=Ret_printIt)
                            b_exit = Button(viewer, text="Exit", bg='orange', width=10, command=Ret_quit)
                            b_print.place(x=120, y=5)
                            b_exit.place(x=300, y=5)
                            pdf = Text(viewer, bg="white")
                            pdf.place(x=0, y=30)
                            # Here the PDF is converted to list of images
                            pages = convert_from_path(ret_entryPass, size=(500, 440), poppler_path=poppler_path)
                            # Empty list for storing images
                            photos = []
                            # Storing the converted images into list
                            for i in range(len(pages)):
                                photos.append(ImageTk.PhotoImage(pages[i]))
                            # Adding all the images to the text widget
                            for photo in photos:
                                pdf.image_create(END, image=photo)
                                # For Seperating the pages
                                pdf.insert(END, '\n\n')
                            # Ending of mainloop
                            mainloop()

                        # Was clicked to generate a new window displaying known details of the viistor and updating which officer is the visitor meeitng today and for what purpose.
                        visitorInfo=tkinter.Toplevel()
                        visitorInfo.title("Visitor Info")
                        visitorInfo.geometry('650x400+250+150')
                        visitorInfo.config(bg='white')
                        # get officer visited, timestamp and purpose of current visit.
                        table = Frame(visitorInfo,bg='white')
                        table.place(x=20, y=30)
                        for i in range(len(labelsForVisitorInfo) - 1):
                            e = Entry(master=table, relief=GROOVE, bg='white')
                            e.grid(row=i, column=0,ipady=5)
                            e.insert(END, labelsForVisitorInfo[i])
                        for j in range(len(returnedVisitorInfo) - 1):
                            e = Entry(master=table, relief=GROOVE, bg='white',width=30)
                            e.grid(row=j, column=1,ipady=5)
                            e.insert(END, returnedVisitorInfo[j])
                        officers = []
                        officerNames=officers_db.find({},{'_id':0,'admin':0})
                        officerCount=officers_db.count_documents({})

                        for i in range(officerCount):
                            officers.append(officerNames[i]["officerName"])
                        global ret_clickedOfficer, ret_purpose_in,ret_areaAllowed
                        areas=["General Area","Licenced Area"]
                        ret_areaAllowed=StringVar()
                        ret_areaAllowed.set("General Area")
                        ret_areaAllowed_in=OptionMenu(table,ret_areaAllowed,*areas)
                        ret_areaAllowed_in.grid(row=8, column=1,ipadx=10)
                        area=Entry(table,bg="white",relief=GROOVE)
                        area.insert(END,'Permission for Area')
                        area.grid(row=8,column=0,ipady=5)

                        ret_clickedOfficer=StringVar()
                        ret_clickedOfficer.set("Officer Name")
                        ret_clickedOfficer_in=OptionMenu(table, ret_clickedOfficer, *officers)
                        ret_clickedOfficer_in.grid(row=6,column=1,ipadx=10)
                        officerToMeet=Entry(table,bg='white', relief=GROOVE)
                        officerToMeet.insert(END,'Officer Name')
                        officerToMeet.grid(row=6,column=0,ipady=5)

                        ret_purpose=Entry(table, bg='white', relief=GROOVE)
                        ret_purpose.insert(END,'Purpose of Visit')
                        ret_purpose_in=Entry(table,width=30)
                        ret_purpose.grid(row=7,column=0,ipady=5)
                        ret_purpose_in.grid(row=7, column=1,ipady=5)

                        visitorImagePath=PhotoImage(file=returnedVisitorInfo[5])
                        visitorImage=Label(visitorInfo, image=visitorImagePath, bg='white')
                        visitorImage.place(x=450,y=80)
                        confirm=Button(visitorInfo,text="Confirm Visitor", command=createPass, width=20, bg='orange')
                        confirm.place(x=380,y=300)
                        mainloop()

                        # Show exisiting information to verify and update purposes
                        # preview and print pass for current visit
                        # update session in db
                        # Clear all values and return to home

                    def clearFindResults():
                        table.place_forget()
                        button_slectVisitor.place_forget()
                        entry_find.delete(0,END)

                    button_slectVisitor=Button(findVisitor, text='Select \n'+returnedVisitorInfo[0]+" "+returnedVisitorInfo[1], width=20,bg='orange', command=button_selectVisitor)
                    button_slectVisitor.place(x=700,y=150)
                    button_clear=Button(findVisitor,text="Clear", width=20, bg='orange', command=clearFindResults)
                    button_clear.place(x=750,y=250)

                createTable()


    home.grid_forget()
    label_find = Label(findVisitor, text="Search by Phone Number", bg='white', fg='black', font=("Verdana", 12))
    # label2=Label(findVisitor, text="Search by First Name")
    entry_find = Entry(findVisitor, width=28, bg="AntiqueWhite1")
    b_find = Button(findVisitor, text='Search', command=searchVisitorButton, bg='orange', fg='black', width=20)
    findVisitor.grid(row=1, sticky='news')
    label_find.place(x=50, y=20)
    entry_find.place(x=286, y=20)
    b_find.place(x=550, y=15)
    button_home = Button(findVisitor, text="Home", width=20, bg='orange', command=click_home)
    button_home.place(x=750, y=15)


# To Open Treeview for retunPass
def returPass():
    hide_all_frames()
    frame_returnPass.grid(row=1,sticky='news')
    footer.grid_forget()

    sqlValues = []
    # sqlValues is a list that needs information in list format for every entry

    def updateDb():
        query = list(sessions.find(
            {"isInside": "yes"},
            {"_id": 0, "firstName": 1, "lastName": 1, "officerVisited": 1, "areaAllowed": 1,"passNo":1}))

        keys = ["firstName", "lastName", "officerVisited", "areaAllowed","passNo"]
        for i in range(len(query)):
            l1 = []
            for j in keys:
                l1.append(query[i][j])
            sqlValues.append(l1)

    # ####### Defing all fucntions

    def updateData():
        updateDb()
        for i in sqlValues:
            # tree.delete()
            tree.insert('', 'end', values=i)

    def closeVisit():
        selected = tree.focus()
        if len(selected) > 0:
            item = tree.item(selected, 'values')
            result1 = messagebox.askyesno("Close Visit", "Do you want to close visit for " + item[0] + " " + item[1])
            if not result1:
                pass
            else:
                closevisit_query = sessions.update_one({'firstName': item[0], "lastName": item[1], "isInside": "yes"},
                                                         {"$set": {'isInside': 'no'}})
                tree.delete(selected)
                countVisitorfn()
        else:
            pass

    ######Declare all variables
    q = StringVar


    # #### LabelFrames
    wrapper1 = Frame(frame_returnPass,bg='white')
    wrapper1.grid(row=0,column=0,padx=40)
    wrapper2 = Frame(frame_returnPass,bg='white')
    wrapper2.grid(row=1,column=0,padx=40)

    ######## Another way of creating a tree
    tree = ttk.Treeview(master=wrapper1, columns=(1, 2, 3, 4,5), show='headings', height='10')
    tree.pack(side='left',padx=35)
    ### adding a scrolbar
    scrollbar=ttk.Scrollbar(wrapper1,orient='vertical',command=tree.yview)
    scrollbar.pack(side='right',fill='y')
    tree.configure(xscrollcommand=scrollbar.set)

    ### adding headings
    headings = ["First Name", "Last Name", "Officer Visited", "Area Visited","Pass ID"]
    # for i in headings:
    #     tree.column((headings.index(i)),width=(len(headings[i])*15))
    a = 0
    while a <= 4:
        tree.heading(a + 1, text=headings[a])
        tree.column(a + 1, width=(len(headings[a]) * 14), minwidth=90, anchor="center")
        a += 1

    updateData()
    def quit_window():
        frame_returnPass.grid_forget()
        click_home()
        footer.grid(row=2, sticky='news')

    def countVisitorfn():
        countVisitor = sessions.count_documents({'isInside': {'$eq': 'yes'}})
        labelVisitorCount.config(text='Visitors Inside: ' + str(countVisitor), font=("Verdana", 10, 'bold'), bg='white',
                                 fg='grey')

    homeButton=Button(wrapper2,text="Home",width=15,bg='orange', fg="black",command=quit_window)
    closeVisitButton = Button(wrapper2, text="Close Visit", width=20, bg='red', fg='white', command=closeVisit)
    homeButton.pack(side='right',padx=20,pady=10)
    closeVisitButton.pack(side='left',padx=20,pady=10)
    labelVisitorCount=Label(frame_returnPass,text="demo")
    labelVisitorCount.grid(row=2,column=0,sticky='ws',padx=70)
    countVisitorfn()


    # tree.pack()

# Header content
logo1=Label(master=header,image=iocl_img, bg='white')
logo1.grid(row=0, column=2)
logo2=Label(master=header,image=atithi_img, bg='white')
logo2.grid(row=0, column=0, ipadx=10)
header1=Label(master=header,text="Visitor Management System", bg="white", font=("Verdana",20,'bold'))
header1.grid(row=0, column=1, ipadx=80)


# Content Content
def generateHomeContent():
    but1=Button(master=home, text='New Visitor', bg="grey", fg="white", width=20, command=click_newVisitor)
    but2=Button(master=home, text='Returning Visitor', bg="grey", fg="white", width=20, command=openFindPage)
    but3=Button(master=home, text='Return Pass', bg="grey", fg="white", width=20, command=returPass)
    but1.place(x=100, y=20)
    but2.place(x=400, y=20)
    but3.place(x=700,y=20)
    home.grid(row=1, sticky='news')
generateHomeContent()

def openSettingsPage():
    # Create a function to check user authentication

    def authenticate_user():
        username = username_entry.get()
        password = password_entry.get()

        # Check if the provided username and password exist in the users collection
        user = users_db.find_one({"username": username, "password": password})

        if user:
            login_window.destroy()  # Close the login window
            create_main_application()
        else:
            messagebox.showerror("Login Failed", "Invalid username or password. Please try again")
            login_window.destroy()

    # Create a login window
    login_window = Tk()
    login_window.title("Login")
    login_window.geometry("300x200+540+170")
    login_window.configure(bg="white")

    # Username label and entry
    username_label = Label(login_window, text="Username:")
    username_label.configure(bg='white')
    username_label.pack(pady=10)
    username_entry = Entry(login_window, width=30)
    username_entry.pack()

    # Password label and entry
    password_label = Label(login_window, text="Password:")
    password_label.configure(bg='white')
    password_label.pack(pady=10)
    password_entry = Entry(login_window, width=30, show="*")
    password_entry.pack()

    # Login button
    login_button = Button(login_window, text="Login", command=authenticate_user, width=20)
    login_button.pack(pady=20)


    # Function to create the main application window
    def create_main_application():

        # Create the main application window
        root = Tk()
        root.title("Settings Page | Atithi 2.1")
        root.geometry("990x600+240+70")
        root.configure(bg="white")

        # Create a notebook (tabbed interface)
        notebook = ttk.Notebook(root)
        notebook.pack(fill='both', expand=True)

        # Create the Officers tab
        officers_tab = ttk.Frame(notebook)
        notebook.add(officers_tab, text="Officers")

        # Create a function to load officer details to the Treeview
        def load_officer_details():
            officers = officers_db.find()
            officers_tree.delete(*officers_tree.get_children())  # Clear existing data
            for officer in officers:
                officers_tree.insert('', 'end', values=(officer['officerName'], officer['admin']))

        # Create a Treeview widget for the officers table
        officers_tree = ttk.Treeview(master=officers_tab, columns=("Name"), height=20)
        officers_tree['show'] = 'headings'
        officers_tree.heading("#1", text="Officer's Name")
        officers_tree.pack()

        # Load officer details on tab load
        load_officer_details()

        # Create a function to add a new officer
        def add_officer():
            new_officer_name = simpledialog.askstring("Add Officer", "Enter the name of the new officer:", parent=root)
            if new_officer_name:
                officers_db.insert_one({"officerName": new_officer_name, "admin": False})
                load_officer_details()

            # Create a function to delete an officer

        def delete_officer():
            selected_items = officers_tree.selection()
            if not selected_items:
                messagebox.showerror("Error", "Please select an officer to delete.")
                return

            item = selected_items[0]  # Take the first selected officer
            officer_name = officers_tree.item(item, 'values')[0]
            officers_db.delete_one({"officerName": officer_name})
            load_officer_details()

        # Create a function to edit officer details
        def edit_officer():
            selected_items = officers_tree.selection()
            if not selected_items:
                messagebox.showerror("Error", "Please select an officer to edit.")
                return

            item = selected_items[0]
            officer_name = officers_tree.item(item, 'values')[0]

            officer = officers_db.find_one({"officerName": officer_name})

            edit_officer_window = Toplevel(root)
            edit_officer_window.title("Edit Officer")
            edit_officer_window.geometry("400x200")

            window_width = edit_officer_window.winfo_reqwidth()
            window_height = edit_officer_window.winfo_reqheight()
            position_x = int((edit_officer_window.winfo_screenwidth() - window_width) / 2)
            position_y = int((edit_officer_window.winfo_screenheight() - window_height) / 2)
            edit_officer_window.geometry(f"+{position_x}+{position_y}")

            edited_officer_name_label = Label(edit_officer_window, text="Officer Name:")
            edited_officer_name_label.pack()
            edited_officer_name = Entry(edit_officer_window, width=50)
            edited_officer_name.insert(0, officer["officerName"])
            edited_officer_name.pack()

            def save_edited_officer_details():
                edited_name = edited_officer_name.get()
                officers_db.update_one(
                    {"officerName": officer_name},
                    {"$set": {"officerName": edited_name}}
                )
                edit_officer_window.destroy()
                load_officer_details()

            save_button = Button(edit_officer_window, text="Save", command=save_edited_officer_details, width=30)
            save_button.pack(pady=10)  # Added pady=10

        # Create a button to edit officers
        button_frame = ttk.Frame(officers_tab)
        add_button = ttk.Button(button_frame, text="Add Officer", command=add_officer, width=30)
        edit_button = ttk.Button(button_frame, text="Edit Officer", command=edit_officer, width=30)
        delete_button = ttk.Button(button_frame, text="Delete Officer", command=delete_officer, width=30)

        add_button.grid(row=0, column=0, padx=5, pady=5)
        edit_button.grid(row=0, column=1, padx=5, pady=5)
        delete_button.grid(row=0, column=2, padx=5, pady=5)
        button_frame.pack()  # Add the button frame to the main officers frame

        # Create the Locations tab
        locations_tab = ttk.Frame(notebook)
        notebook.add(locations_tab, text="Locations")

        # Create a function to load location details to the Treeview
        # def load_location_details():
        #     locations = location_db.find()
        #     locations_tree.delete(*locations_tree.get_children())  # Clear existing data
        #     for location in locations:
        #         locations_tree.insert('', 'end', values=(
        #             location['location'], location['locationType'], location['address'],
        #             location['admin'], location['allowedVisitors']))
        #
        # # Create a Treeview widget for the locations table
        # locations_tree = ttk.Treeview(locations_tab,
        #                               columns=("Location", "Type", "Address", "Admin", "Allowed Visitors"))
        # locations_tree['show'] = 'headings'
        # locations_tree.heading("#1", text="Location")
        # locations_tree.heading("#2", text="Type")
        # locations_tree.heading("#3", text="Address")
        # locations_tree.heading("#4", text="Admin")
        # locations_tree.heading("#5", text="Allowed Visitors")
        # locations_tree.pack()
        #
        # # Load location details on tab load
        # load_location_details()
        #
        # # Create a function to edit location details
        # def edit_location():
        #     selected_items = locations_tree.selection()
        #     if not selected_items:
        #         messagebox.showerror("Error", "Please select a location to edit.")
        #         return
        #
        #     item = selected_items[0]
        #     location_name = locations_tree.item(item, 'values')[0]
        #
        #     location = location_db.find_one({"location": location_name})
        #
        #     edit_location_window = Toplevel(root)
        #     edit_location_window.title("Edit Location")
        #     edit_location_window.geometry("600x400")
        #
        #     window_width = edit_location_window.winfo_reqwidth()
        #     window_height = edit_location_window.winfo_reqheight()
        #     position_x = int((edit_location_window.winfo_screenwidth() - window_width) / 2)
        #     position_y = int((edit_location_window.winfo_screenheight() - window_height) / 2)
        #     edit_location_window.geometry(f"+{position_x}+{position_y}")
        #
        #     edited_location_name_label = Label(edit_location_window, text="Location:")
        #     edited_location_name_label.pack()
        #     edited_location_name = Entry(edit_location_window, width=50)
        #     edited_location_name.insert(0, location["location"])
        #     edited_location_name.pack()
        #
        #     edited_location_type_label = Label(edit_location_window, text="Location Type:")
        #     edited_location_type_label.pack()
        #     edited_location_type = Entry(edit_location_window, width=50)
        #     edited_location_type.insert(0, location["locationType"])
        #     edited_location_type.pack()
        #
        #     edited_location_address_label = Label(edit_location_window, text="Address:")
        #     edited_location_address_label.pack()
        #     edited_location_address = Entry(edit_location_window, width=50)
        #     edited_location_address.insert(0, location["address"])
        #     edited_location_address.pack()
        #
        #     edited_location_admin_label = Label(edit_location_window, text="Admin:")
        #     edited_location_admin_label.pack()
        #     edited_location_admin = Entry(edit_location_window, width=50)
        #     edited_location_admin.insert(0, location["admin"])
        #     edited_location_admin.pack()
        #
        #     edited_location_allowed_label = Label(edit_location_window, text="Allowed Visitors:")
        #     edited_location_allowed_label.pack()
        #     edited_location_allowed = Entry(edit_location_window, width=50)
        #     edited_location_allowed.insert(0, location["allowedVisitors"])
        #     edited_location_allowed.pack()
        #
        #     def save_edited_location_details():
        #         edited_name = edited_location_name.get()
        #         edited_type = edited_location_type.get()
        #         edited_address = edited_location_address.get()
        #         edited_admin = edited_location_admin.get()
        #         edited_allowed = edited_location_allowed.get()
        #
        #         location_db.update_one(
        #             {"location": location_name},
        #             {
        #                 "$set": {
        #                     "location": edited_name,
        #                     "locationType": edited_type,
        #                     "address": edited_address,
        #                     "admin": edited_admin,
        #                     "allowedVisitors": edited_allowed,
        #                 }
        #             }
        #         )
        #         edit_location_window.destroy()
        #         load_location_details()
        #
        #     save_button = Button(edit_location_window, text="Save", command=save_edited_location_details, width=40)
        #     save_button.pack(pady=10)  # Added pady=10
        #
        # # Create a button to edit locations
        # edit_location_button = Button(locations_tab, text="Edit Location", command=edit_location)
        # edit_location_button.pack(pady=10)  # Added pady=10

        root.mainloop()

    login_window.mainloop()


def click_report_button():

    # Create the reports application window
    report_root = Tk()
    report_root.title("Report Generator")
    report_root.geometry('990x600+240+70')
    report_root.configure(bg="white")

    # Creating a wrapper for adding the treeview results.
    wrapper_results = Frame(report_root, bg="white")
    wrapper_results.grid(row=0, column=0)

    wrapper_results2 = Frame(report_root, bg="white")
    wrapper_results2.grid(row=1, column=0)

    # Create a custom font for buttons
    # button_font = tkfont.Font(family='Helvetica', size=12, weight='bold')

    # Filter criteria dictionary
    filter_criteria = {}

    # Function to format datetime.date objects to string
    def format_date_to_string(date_obj):
        return date_obj.strftime("%d/%m/%Y")

    # Function to apply filters and generate a report
    def generate_report():

        # results = collection.find({'entryDate': {'$gte': date_filter}})
        # Clear the previous report data
        for record in results_tree.get_children():
            results_tree.delete(record)

        ### defining variables that were enterd in the search query
        visitor_name = visitor_name_entry.get().strip()
        officer_name = officer_name_combobox.get().strip()

        # Date Range filter
        date_range = date_range_combobox.get()
        if date_range == "Last 7 Days" or not date_range:  # Handle the "Last 7 Days" and default case
            start_date_str = format_date_to_string(datetime.now() - timedelta(days=7))
            start_date_str = datetime.strptime(start_date_str, "%d/%m/%Y")
            date_filter = {"$gte": start_date_str}
        elif date_range == "Custom Date Range":
            try:
                start_date = custom_start_date_calendar.get_date()
                end_date = custom_end_date_calendar.get_date() + timedelta(days=1)
                start_date_str = format_date_to_string(start_date)
                start_date_str = datetime.strptime(start_date_str, "%d/%m/%Y")
                end_date_str = format_date_to_string(end_date)
                end_date_str = datetime.strptime(end_date_str, "%d/%m/%Y")
                date_filter = {"$gte": start_date_str, "$lt": end_date_str}
            except ValueError:
                pass

        if not visitor_name and not officer_name:
            filter_criteria["isoDate"] = date_filter

        # Visitor Name filter
        if visitor_name:
            # filter_criteria['$or'] = [{'firstName': {'$regex': visitor_name, '$options': 'i'}},
            #                           {'lastName': {'$regex': visitor_name, '$options': 'i'}}]
            filter_criteria['$and'] = [
                {
                    '$or': [
                        {'firstName': {'$regex': visitor_name, '$options': 'i'}},
                        {'lastName': {'$regex': visitor_name, '$options': 'i'}}
                    ]
                },
                {'isoDate': date_filter}
            ]

        # Officer Name filter
        if officer_name:
            # filter_criteria['officerVisited'] = officer_name
            filter_criteria['$and'] = [{'officerVisited': officer_name,
                                        'isoDate': date_filter}]

        ## when both conditions are given
        if visitor_name and officer_name:
            filter_criteria["$and"] = [{
                '$or': [
                    {'firstName': {'$regex': visitor_name, '$options': 'i'}},
                    {'lastName': {'$regex': visitor_name, '$options': 'i'}}
                ]
            },
                {'officerVisited': officer_name},
                {'isoDate': date_filter}
            ]

        # Query the database with filter criteria
        results = sessions.find(filter_criteria)
        # print(results[0])

        # Populate the treeview with results
        for result in results:
            results_tree.insert('', 'end', values=(result['sessionId'], result['firstName'], result['lastName'],
                                                   result['entryDateTime'], result['officerVisited'],
                                                   result['areaAllowed'], result['passNo']))

        # Clear the filter criteria for the next query
        filter_criteria.clear()

    # Function to reset filters and clear the treeview
    def reset_filters():
        visitor_name_entry.delete(0, END)
        date_range_combobox.set("")
        custom_start_date_calendar.set_date(datetime.now() - timedelta(days=7))
        custom_end_date_calendar.set_date(datetime.now())
        officer_name_combobox.set("")
        for record in results_tree.get_children():
            results_tree.delete(record)

    # Function to export report to Excel
    def export_to_excel():
        if not results_tree.get_children():
            messagebox.showerror("Export Error", "No data to export.")
            return

        file_location = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                     filetypes=[("Excel Files", "*.xlsx")],
                                                     initialdir="~/Desktop/")
        if not file_location:
            return

        wb = Workbook()
        ws = wb.active

        columns = ["Session ID", "First Name", "Last Name", "Entry Date/Time",
                   "Officer Visited", "Area Allowed", "Pass No"]

        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal="center")

        for row_num, item in enumerate(results_tree.get_children(), 2):
            for col_num, value in enumerate(results_tree.item(item)['values'], 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

        try:
            wb.save(file_location)
            messagebox.showinfo("Export Success", "Report exported to Excel successfully.")
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting to Excel: {str(e)}")

    # Function to export report to PDF
    # def export_to_pdf():
    #     if not results_tree.get_children():
    #         messagebox.showerror("Export Error", "No data to export.")
    #         return
    #
    #     file_location = filedialog.asksaveasfilename(defaultextension=".pdf",
    #                                                 filetypes=[("PDF Files", "*.pdf")],
    #                                                 initialdir="~/Desktop/")
    #     if not file_location:
    #         return
    #
    #     doc = SimpleDocTemplate(file_location, pagesize=letter)
    #     data = []
    #     data.append(["Session ID", "First Name", "Last Name", "Entry Date/Time",
    #                  "Officer Visited", "Area Allowed", "Pass No"])
    #     for item in results_tree.get_children():
    #         data.append(results_tree.item(item)['values'])
    #
    #     table = Table(data)
    #     style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    #                         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    #                         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    #                         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    #                         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    #                         ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    #                         ('GRID', (0, 0), (-1, -1), 1, colors.black)])
    #
    #     table.setStyle(style)
    #     data.append(table)
    #
    #     doc.build(data)
    #     messagebox.showinfo("Export Success", "Report exported to PDF successfully.")

    # Search Criteria
    # Visitor Name Label and Entry
    visitor_name_label = Label(wrapper_results, text="Visitor Name:", bg="white")
    visitor_name_label.grid(row=0, column=0, padx=10, pady=10)
    visitor_name_entry = Entry(wrapper_results)
    visitor_name_entry.grid(row=0, column=1, padx=10, pady=10)

    # Date Range Label and Combobox
    date_range_label = Label(wrapper_results, text="Date Range:", bg="white")
    date_range_label.grid(row=0, column=2, padx=10, pady=10)
    date_ranges = ["Last 7 Days", "Custom Date Range"]
    date_range_combobox = ttk.Combobox(wrapper_results, values=date_ranges)
    date_range_combobox.grid(row=0, column=3, padx=10, pady=10)
    date_range_combobox.set("Last 7 Days")

    # Custom Date Range Label and Calendar Widgets
    custom_start_date_label = Label(wrapper_results, text="Start Date:", bg="white")
    custom_start_date_label.grid(row=0, column=4, padx=10, pady=10)
    custom_start_date_calendar = tkcalendar.DateEntry(wrapper_results, width=12, background='orange',
                                                      foreground='white',
                                                      borderwidth=2)
    custom_start_date_calendar.grid(row=0, column=5, padx=10, pady=10)
    custom_start_date_calendar.set_date(datetime.now() - timedelta(days=7))

    custom_end_date_label = Label(wrapper_results, text="End Date:", bg="white")
    custom_end_date_label.grid(row=0, column=6, padx=10, pady=10)
    custom_end_date_calendar = tkcalendar.DateEntry(wrapper_results, width=12, background='darkblue',
                                                    foreground='white',
                                                    borderwidth=2)
    custom_end_date_calendar.grid(row=0, column=7, padx=10, pady=10)
    custom_end_date_calendar.set_date(datetime.now())

    # Officer Name Label and Combobox
    officer_name_label = Label(wrapper_results, text="Officer Name:", bg="white")
    officer_name_label.grid(row=1, column=0, padx=10, pady=10)
    officers = list(set([officer['officerVisited'] for officer in sessions.find()]))
    officer_name_combobox = ttk.Combobox(wrapper_results, values=officers)
    officer_name_combobox.grid(row=1, column=1, padx=10, pady=10)
    officer_name_combobox.set("")

    # Generate Report Button
    generate_report_button = Button(wrapper_results, text="Generate Report", command=generate_report, bg="orange")
    generate_report_button.grid(row=1, column=2, padx=10, pady=10, columnspan=2)

    # Reset Filters Button
    reset_filters_button = Button(wrapper_results, text="Reset Filters", command=reset_filters, bg="orange")
    reset_filters_button.grid(row=1, column=4, padx=10, pady=10)

    # Export to Excel Button
    export_excel_button = Button(wrapper_results, text="Export to Excel", command=export_to_excel, bg="orange")
    export_excel_button.grid(row=1, column=5, padx=(20, 10), pady=10)

    # # Export to PDF Button
    # export_pdf_button = Button(report_root, text="Export to PDF", command=export_to_pdf, font=button_font, bg="orange")
    # export_pdf_button.grid(row=1, column=6, padx=10, pady=10)

    # Create Treeview for Results
    columns = ["Session ID", "First Name", "Last Name", "Entry Date/Time",
               "Officer Visited", "Area Allowed", "Pass No"]

    results_tree = ttk.Treeview(wrapper_results2, columns=columns, show="headings", selectmode="browse", height=20)
    results_tree.grid(row=2, column=0, padx=10, pady=10, columnspan=8)

    ### adding a scrolbar
    scrollbar = ttk.Scrollbar(wrapper_results2, orient='vertical', command=results_tree.yview)
    scrollbar.grid(row=2, column=8, padx=2, pady=10, rowspan=20)
    results_tree.configure(xscrollcommand=scrollbar.set)

    for col in columns:
        results_tree.heading(col, text=col)
        results_tree.column(col, width=100)

    # Main loop
    report_root.mainloop()

#New Visitor content
headernV = Label(master=newVisitor, text="New Visitor Registration", bg="white", font=("Verdana", 12, "bold"))

# # Drop Down cases
# Govt ID
govtIDoptions = ["Aadhar Card", "Voter Card", "Driver's Licence", "Other"]
clickedId = StringVar()
clickedId.set("Govt Photo ID")
photoId = OptionMenu(newVisitor, clickedId, *govtIDoptions)
# Area Allowed
areas = ["General Area", "Licenced Area"]
areaAllowed = StringVar()
areaAllowed.set("General Area")
areaAllowed_in = OptionMenu(newVisitor, areaAllowed, *areas)
# Officers
query_officer=officers_db.find({},{"_id":0, "admin":0})
count_officers=officers_db.count_documents({})
officers = []
for i in range(count_officers):
    officers.append(query_officer[i]["officerName"])

clickedOfficers = StringVar()
clickedOfficers.set("Officer Name")
officerVisited_in = OptionMenu(newVisitor, clickedOfficers, *officers)

# # #  Other labels
firstName = Label(newVisitor, text='First Name', bg='white', font=("Verdana", 12, 'bold'))
lastName = Label(newVisitor, text='Last Name', bg='white', font=("Verdana", 12, 'bold'))
mobileNumber = Label(newVisitor, text='Mobile', bg='white', font=("Verdana", 12, 'bold'))
address = Label(newVisitor, text='Full Address', bg='white', font=("Verdana", 12, 'bold'))
area=Label(newVisitor,text="Permission for area", bg='white', font=("Verdana", 12, 'bold'))
officerVisited = Label(newVisitor, text='Officer To Be Visited', bg='white', font=("Verdana", 12, 'bold'))
purpose = Label(newVisitor, text='Purpose', bg='white', font=("Verdana", 12, 'bold'))

# # # Entries
firstName_in=Entry(newVisitor, width=20)
lastName_in=Entry(newVisitor, width=20)
mobileNumber_in=Entry(newVisitor, width=20)
photoId_in=Entry(newVisitor, width=20)
address_in=Entry(newVisitor, width=20)

purpose_in = Entry(newVisitor, width=20)
reset = Button(newVisitor, bg='orange', fg="black", text='Reset', width='20', command=click_reset)
submit = Button(newVisitor, bg='orange', fg="black", text='Submit', width='20', command=submit)
b_home=Button(newVisitor, bg='orange', fg="black", text='Home', width='20', command=click_home)
b_clickPicture=Button(newVisitor, bg='orange', fg='black', width=20, text="Click Picture", command=clickPicture)


# Search existing visitors
### Other features
#settings page.


# Footer content
timer=Label(master=footer, bg='black', fg='white', justify='right')
timer.pack(side=RIGHT, padx=13)

setting=Label(master=footer, bg='black', fg='orange', text="Settings")
setting.bind("<Button-1>", lambda e:(openSettingsPage()))
# <Old setting prompt>
# setting.bind("<Button-1>", lambda e:(messagebox.showerror("Error", "Settings are disabled in this subsription. Please contact Egis Engineering")))
setting.place(x=10, y=15)
support=Label(master=footer, bg='black', fg='orange', text="Support")
support.bind("<Button-1>", lambda e:(messagebox.showinfo("Contact Us","Please visit www.egisengineering.in for support")))
support.place(x=10, y=45)
reports=Label(master=footer, bg='black', fg='orange', text="Reports")
reports.bind("<Button-1>", lambda e:(click_report_button()))
reports.place(x=130, y=15)
exited=Label(master=footer, bg='black', fg='orange', text="Exit")
exited.bind("<Button-1>", lambda e:(window.destroy()))
exited.place(x=130, y=45)

locationType_query=str(location_db.find({},{"locationType":1,"_id":0})[0]["locationType"])
address_query=str(location_db.find({},{"address":1,"_id":0})[0]["address"])
admin_query=str(location_db.find({},{"admin":1,"_id":0})[0]["admin"])

location=Label(master=footer, bg='black', fg='white', text=locationType_query+", "+address_query+'\n Officer Incharge: '+admin_query)
location.place(x=400, y=35)

# Settings_for_grid_tkinter
window.columnconfigure(0, weight=1) # 100%

window.rowconfigure(0, weight=1) # 10%
window.rowconfigure(1, weight=8) # 80%
window.rowconfigure(2, weight=1) # 10%

header.grid(row=0, sticky='news')
footer.grid(row=2, sticky='news')


# Functions post load
def digitalclock():
   text_input = time.strftime("%A, %d/%B/%Y\nTime: %H:%M:%S")
   timer.config(text=text_input)
   timer.after(200, digitalclock)

digitalclock()

window.mainloop()
